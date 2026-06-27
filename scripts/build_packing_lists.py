#!/usr/bin/env python3
from __future__ import annotations

import json
import os
import re
import signal
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote

from pypdf import PdfReader

from extract_model_components import (
    ACCESSORY_KEYWORDS,
    COMPONENT_KEYWORDS,
    STOP_KEYWORDS,
    extract_list,
    find_section,
    models_from_filename,
    short_context,
)


ROOT = Path(__file__).resolve().parents[1]
DATA_PATH = ROOT / "knowledge-base" / "assets" / "manuals-data.js"
MODEL_TYPE_PATH = ROOT / "outputs" / "model_type_review" / "型号机器类型备注表.xlsx"
MODEL_TYPE_JSON_PATH = ROOT / "scripts" / "model_type_overrides.json"
OUT_DIR = ROOT / "knowledge-base" / "assets"
PACKING_DIR = OUT_DIR / "packing"
PACKING_DATA_PATH = OUT_DIR / "packing-data.js"
TMP_DIR = ROOT / "tmp" / "pdfs" / "packing"
VISION_OCR_SOURCE = ROOT / "scripts" / "vision_ocr.swift"

PAGE_SCAN_LIMIT = 24
OCR_SCAN_LIMIT = 16
SCREENSHOT_SCALE = "1200"
OCR_SCALE = "1100"
PROCESS_MACHINE_TYPES = {
    "反渗透净水机",
    "前置过滤器",
    "超滤机",
    "净热一体机",
    "管线机",
    "商用直饮机",
}

PAGE_KEYWORDS = [
    ("装箱清单如下", 90),
    ("装箱清单", 86),
    ("包装清单", 84),
    ("随箱附件", 82),
    ("附件清单", 80),
    ("产品配件", 76),
    ("产品组件", 70),
    ("产品部件", 62),
    ("各部件", 56),
    ("组件名称", 54),
    ("纸箱内还包括", 82),
    ("包装箱内", 66),
    ("彩盒内", 66),
    ("配件包", 58),
    ("组件包", 56),
    ("开箱", 48),
]
EXPLICIT_PAGE_KEYWORDS = {
    "装箱清单如下",
    "装箱清单",
    "包装清单",
    "随箱附件",
    "附件清单",
    "产品配件",
    "产品组件",
    "产品部件",
    "各部件",
    "组件名称",
}
CONTAINER_PAGE_KEYWORDS = {
    "纸箱内还包括",
    "包装箱内",
    "彩盒内",
}
GENERIC_PAGE_KEYWORDS = {
    "配件包",
    "组件包",
    "开箱",
}
INSTALL_PAGE_TERMS = [
    "安装提示",
    "安装步骤",
    "安装前的准备工作",
    "安装水龙头",
    "部件连接",
    "连接净水端",
    "连接管路",
    "水管安装",
    "首次安装",
]

POSITIVE_TERMS = [
    "主机",
    "说明书",
    "用户手册",
    "附件",
    "配件",
    "电源",
    "适配器",
    "水龙头",
    "零件",
    "滤芯",
    "遥控器",
    "数量",
    "名称",
    "包装",
    "箱内",
    "包括",
    "包含",
    "一套",
    "一只",
    "一个",
    "1套",
    "1只",
    "1个",
]

NEGATIVE_TERMS = [
    "目录",
    "有害物质",
    "环保清单",
    "质保",
    "保修",
    "故障",
    "滤芯更换",
    "水路图",
    "电气原理",
]


class PageExtractTimeout(Exception):
    pass


def _raise_timeout(_signum, _frame) -> None:
    raise PageExtractTimeout()


def bundled_tool(name: str) -> str:
    env_name = name.upper().replace("-", "_")
    if os.environ.get(env_name):
        return os.environ[env_name]
    local = shutil.which(name)
    if local:
        return local
    bundled = (
        Path.home()
        / ".cache/codex-runtimes/codex-primary-runtime/dependencies/bin"
        / name
    )
    if bundled.exists():
        return str(bundled)
    raise FileNotFoundError(f"Missing required tool: {name}")


def load_manuals() -> list[dict]:
    text = DATA_PATH.read_text(encoding="utf-8")
    match = re.search(r"window\.MANUALS\s*=\s*(\[.*\]);\s*$", text, re.S)
    if not match:
        raise ValueError(f"Could not find MANUALS data in {DATA_PATH}")
    return json.loads(match.group(1))


def load_model_types() -> dict[str, str]:
    if MODEL_TYPE_PATH.exists():
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("Missing openpyxl; cannot read model type workbook") from exc

        workbook = openpyxl.load_workbook(MODEL_TYPE_PATH, data_only=True)
        sheet = workbook["型号备注表"]
        headers = [cell.value for cell in sheet[1]]
        model_col = headers.index("型号") + 1
        type_col = headers.index("你备注的机器类型") + 1

        mapping: dict[str, str] = {}
        for row in range(2, sheet.max_row + 1):
            model = sheet.cell(row, model_col).value
            machine_type = sheet.cell(row, type_col).value
            if not model:
                continue
            mapping[str(model).strip().upper()] = str(machine_type or "").strip()
        return mapping

    if MODEL_TYPE_JSON_PATH.exists():
        payload = json.loads(MODEL_TYPE_JSON_PATH.read_text(encoding="utf-8"))
        return {
            str(model).strip().upper(): str(machine_type or "").strip()
            for model, machine_type in payload.items()
        }

    raise FileNotFoundError(
        f"Missing model type data: {MODEL_TYPE_PATH} or {MODEL_TYPE_JSON_PATH}"
    )


def clean_text(value: str) -> str:
    value = (value or "").replace("\u3000", " ")
    value = re.sub(r"[ \t\r\f\v]+", " ", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def compact_source(value: str, max_len: int = 900) -> str:
    value = clean_text(value)
    value = re.sub(r"\s+", " ", value)
    return value[:max_len].rstrip()


def pdf_path_for(manual: dict) -> Path:
    file_url = unquote(manual.get("fileUrl") or "")
    return (ROOT / "knowledge-base" / file_url).resolve()


def lookup_model_type(model: str, manual: dict, model_types: dict[str, str]) -> str:
    key = model.upper()
    if key in model_types:
        return model_types[key]

    hay = " ".join([manual.get("title") or "", manual.get("filename") or ""]).upper()
    compact_hay = re.sub(r"[^A-Z0-9]+", "", hay)
    compact_key = re.sub(r"[^A-Z0-9]+", "", key)

    title_matches = []
    for model_key, machine_type in model_types.items():
        compact_model = re.sub(r"[^A-Z0-9]+", "", model_key)
        if not compact_model or len(compact_model) < 4 or not machine_type:
            continue
        if compact_model in compact_hay:
            title_matches.append((len(compact_model), machine_type))
    if title_matches:
        title_matches.sort(reverse=True)
        return title_matches[0][1]

    compact_matches = []
    for model_key, machine_type in model_types.items():
        compact_model = re.sub(r"[^A-Z0-9]+", "", model_key)
        if not compact_model or not machine_type:
            continue
        if compact_model == compact_key:
            return machine_type
        if (
            compact_key
            and compact_model.startswith(compact_key)
            and compact_model in compact_hay
        ):
            compact_matches.append((len(compact_model), machine_type))
    if compact_matches:
        compact_matches.sort(reverse=True)
        return compact_matches[0][1]

    if key.isdigit():
        for suffix in ("VC", "WHF", "ECM"):
            alias = f"{key}{suffix}"
            if alias in compact_hay and alias in model_types:
                return model_types[alias]
    return ""


def manual_model_types(manual: dict, model_types: dict[str, str]) -> dict[str, str]:
    models = models_from_filename(manual.get("filename", ""), manual.get("models") or [])
    return {model: lookup_model_type(model, manual, model_types) for model in models}


def skip_reason(manual: dict, model_types: dict[str, str]) -> str:
    typed_models = manual_model_types(manual, model_types)
    types = {value for value in typed_models.values() if value}
    if types and types.isdisjoint(PROCESS_MACHINE_TYPES):
        return "、".join(sorted(types))
    return ""


def compile_ocr_tool() -> Path | None:
    swiftc = shutil.which("swiftc")
    if not swiftc or not VISION_OCR_SOURCE.exists():
        return None

    TMP_DIR.mkdir(parents=True, exist_ok=True)
    tool_path = TMP_DIR / "vision_ocr"
    if (
        tool_path.exists()
        and tool_path.stat().st_mtime >= VISION_OCR_SOURCE.stat().st_mtime
    ):
        return tool_path

    subprocess.run(
        [swiftc, str(VISION_OCR_SOURCE), "-o", str(tool_path)],
        check=True,
        capture_output=True,
        text=True,
        timeout=60,
    )
    return tool_path


def extract_page_texts(pdf_path: Path, max_pages: int) -> list[dict]:
    try:
        reader = PdfReader(str(pdf_path))
    except Exception:
        return []

    texts: list[dict] = []
    limit = min(len(reader.pages), max_pages)
    for index in range(limit):
        previous_handler = signal.signal(signal.SIGALRM, _raise_timeout)
        signal.alarm(3)
        try:
            text = reader.pages[index].extract_text() or ""
        except (Exception, PageExtractTimeout):
            text = ""
        finally:
            signal.alarm(0)
            signal.signal(signal.SIGALRM, previous_handler)
        texts.append({"page": index + 1, "text": clean_text(text), "source": "pdf-text"})
    return texts


def page_score(text: str) -> tuple[int, str]:
    normalized = re.sub(r"\s+", "", text or "")
    if not normalized:
        return (0, "")

    score = 0
    best_keyword = ""
    for keyword, weight in PAGE_KEYWORDS:
        if keyword in normalized:
            score += weight
            if not best_keyword:
                best_keyword = keyword

    for term in POSITIVE_TERMS:
        if term in normalized:
            score += 5

    for term in NEGATIVE_TERMS:
        if term in normalized:
            score -= 22

    if best_keyword in GENERIC_PAGE_KEYWORDS and any(
        term in normalized for term in INSTALL_PAGE_TERMS
    ):
        score -= 60
    elif best_keyword in CONTAINER_PAGE_KEYWORDS and any(
        term in normalized for term in INSTALL_PAGE_TERMS
    ):
        score -= 35
    if "目录" in normalized[:120]:
        score -= 65
    if re.search(r"序号.*名称.*数量", normalized):
        score += 28
    if "开箱" in normalized and not re.search(
        r"纸箱内还包括|箱内|包装箱|彩盒|配件|附件|小零件|零件包", normalized
    ):
        score -= 38
    if re.search(r"(?:\d+|[一二三四五六七八九十]+)[\.、）)]", normalized):
        score += 8

    return score, best_keyword


def page_keyword_tier(keyword: str) -> int:
    if keyword in EXPLICIT_PAGE_KEYWORDS:
        return 3
    if keyword in CONTAINER_PAGE_KEYWORDS:
        return 2
    if keyword in GENERIC_PAGE_KEYWORDS:
        return 1
    return 0


def page_candidate_key(candidate: dict) -> tuple[int, int, int, int]:
    source_priority = 1 if candidate.get("source") == "ocr" else 0
    return (
        page_keyword_tier(candidate.get("keyword") or ""),
        candidate.get("score") or 0,
        source_priority,
        -int(candidate.get("page") or 0),
    )


def best_page(pages: list[dict], min_score: int = 42) -> dict | None:
    best: dict | None = None
    for page in pages:
        score, keyword = page_score(page.get("text", ""))
        if score < min_score:
            continue
        candidate = {
            "page": page["page"],
            "text": page.get("text", ""),
            "source": page.get("source", ""),
            "score": score,
            "keyword": keyword,
            "image_path": page.get("image_path", ""),
        }
        if best is None or page_candidate_key(candidate) > page_candidate_key(best):
            best = candidate
    return best


def image_page_number(path: Path) -> int:
    match = re.search(r"-(\d+)\.png$", path.name)
    return int(match.group(1)) if match else 0


def ocr_front_pages(
    pdf_path: Path,
    pdftoppm: str,
    ocr_tool: Path | None,
    max_pages: int,
) -> list[dict]:
    if not ocr_tool:
        return []

    with tempfile.TemporaryDirectory(prefix="scan_", dir=TMP_DIR) as temp_dir:
        temp_path = Path(temp_dir)
        prefix = temp_path / "page"
        try:
            subprocess.run(
                [
                    pdftoppm,
                    "-f",
                    "1",
                    "-l",
                    str(max_pages),
                    "-png",
                    "-scale-to",
                    OCR_SCALE,
                    str(pdf_path),
                    str(prefix),
                ],
                check=True,
                capture_output=True,
                timeout=90,
            )
        except Exception:
            return []

        images = sorted(temp_path.glob("page-*.png"), key=image_page_number)
        if not images:
            return []

        try:
            result = subprocess.run(
                [str(ocr_tool), *map(str, images)],
                check=True,
                capture_output=True,
                text=True,
                timeout=max(45, 8 * len(images)),
            )
            payload = json.loads(result.stdout)
        except Exception:
            return []

        pages = []
        for image, item in zip(images, payload):
            pages.append(
                {
                    "page": image_page_number(image),
                    "text": clean_text(item.get("text") or ""),
                    "source": "ocr",
                    "image_path": str(image),
                }
            )
        return pages


def render_screenshot(
    pdf_path: Path,
    manual_id: str,
    page: int,
    pdftoppm: str,
    existing_image: str = "",
) -> str:
    PACKING_DIR.mkdir(parents=True, exist_ok=True)
    safe_id = re.sub(r"[^A-Za-z0-9_-]+", "-", str(manual_id))
    out_prefix = PACKING_DIR / f"manual-{safe_id}-p{page}"
    png_path = out_prefix.with_suffix(".png")
    rel_path = f"assets/packing/{png_path.name}"

    if png_path.exists():
        return rel_path

    if existing_image:
        try:
            shutil.copyfile(existing_image, png_path)
            return rel_path
        except Exception:
            pass

    try:
        subprocess.run(
            [
                pdftoppm,
                "-f",
                str(page),
                "-l",
                str(page),
                "-singlefile",
                "-png",
                "-scale-to",
                SCREENSHOT_SCALE,
                str(pdf_path),
                str(out_prefix),
            ],
            check=True,
            capture_output=True,
            timeout=45,
        )
    except Exception:
        return ""
    return rel_path if png_path.exists() else ""


def sections_from_text(text: str) -> dict:
    component_snippet, component_keyword, component_status = find_section(
        text, COMPONENT_KEYWORDS, prefer_list=False
    )
    accessory_snippet, accessory_keyword, accessory_status = find_section(
        text, ACCESSORY_KEYWORDS, prefer_list=True
    )

    components = extract_list(component_snippet)
    accessories = extract_list(accessory_snippet)
    if not accessories and component_keyword and "装箱清单" in component_snippet:
        accessories = components
        accessory_keyword = component_keyword
        accessory_status = "found_in_component_section"

    if not components:
        component_status = "not_found"
    if not accessories:
        accessory_status = "not_found"

    source_parts = []
    if component_snippet:
        source_parts.append(f"产品组件来源：{short_context(component_snippet, 420)}")
    if accessory_snippet:
        source_parts.append(f"装箱清单来源：{short_context(accessory_snippet, 520)}")

    found = bool(components or accessories)
    return {
        "status": "found" if found else "not_found",
        "components": components,
        "accessories": accessories,
        "componentStatus": component_status,
        "accessoryStatus": accessory_status,
        "componentKeyword": component_keyword,
        "accessoryKeyword": accessory_keyword,
        "sourceText": "\n\n".join(source_parts),
        "keyword": accessory_keyword or component_keyword,
    }


def manual_sections(manual: dict) -> dict:
    return sections_from_text(manual.get("text") or "")


def weak_extracted_list(value: str) -> bool:
    lines = [line.strip() for line in (value or "").splitlines() if line.strip()]
    if not lines or len(lines) > 8:
        return False
    normalized = " ".join(lines)
    toc_terms = [*STOP_KEYWORDS, "功能示意图", "滤芯名称", "目录"]
    stop_hits = sum(1 for keyword in toc_terms if keyword in normalized)
    if "功能示意图" in normalized and "滤芯名称" in normalized:
        return True
    item_hits = sum(
        1
        for keyword in ("主机", "滤芯", "水龙头", "说明书", "用户手册", "配件", "接头", "PE管")
        if keyword in normalized
    )
    return stop_hits >= 2 and item_hits == 0


def noisy_extracted_list(value: str) -> bool:
    lines = [line.strip() for line in (value or "").splitlines() if line.strip()]
    if not lines:
        return False
    head = " ".join(lines[:4])
    if any(
        term in head
        for term in (
            "产品功能",
            "功能示意图",
            "滤芯名称",
            "技术参数",
            "安装提示",
            "产品介绍",
            "产品维修指南",
            "产品保养",
            "质保条款",
            "外观示意图",
            "故障代码",
        )
    ):
        return True
    normalized = " ".join(lines)
    if any(
        term in normalized
        for term in ("OWNER'S", "MANUAL", "扫一扫", "净享生活", "顶盖上方", "市政自来水")
    ):
        return True
    if re.search(r"\d{2,4}\s*mm", normalized, re.I):
        return True
    return weak_extracted_list(value)


def sections_useful(sections: dict) -> bool:
    values = [sections.get("components") or "", sections.get("accessories") or ""]
    useful = [value for value in values if value and not noisy_extracted_list(value)]
    return bool(useful)


def empty_sections() -> dict:
    return {
        "status": "not_found",
        "components": "",
        "accessories": "",
        "componentStatus": "not_found",
        "accessoryStatus": "not_found",
        "componentKeyword": "",
        "accessoryKeyword": "",
        "sourceText": "",
        "keyword": "",
    }


def candidate_needs_ocr(candidate: dict | None, sections: dict) -> bool:
    if not candidate:
        return True
    keyword = candidate.get("keyword") or ""
    if keyword not in EXPLICIT_PAGE_KEYWORDS:
        return True
    if candidate.get("source") == "pdf-text" and not sections_useful(sections):
        return True
    return False


def choose_best_with_ocr(pdf_pages: list[dict], ocr_pages: list[dict]) -> dict | None:
    combined_candidate = best_page([*pdf_pages, *ocr_pages])
    ocr_candidate = best_page(ocr_pages)
    if not combined_candidate or not ocr_candidate:
        return combined_candidate

    combined_sections = sections_from_text(combined_candidate.get("text") or "")
    ocr_sections = sections_from_text(ocr_candidate.get("text") or "")
    if sections_useful(ocr_sections) and not sections_useful(combined_sections):
        return ocr_candidate
    if (
        sections_useful(ocr_sections)
        and ocr_candidate.get("page") == combined_candidate.get("page")
        and page_keyword_tier(ocr_candidate.get("keyword") or "")
        >= page_keyword_tier(combined_candidate.get("keyword") or "")
    ):
        return ocr_candidate
    return combined_candidate


def merge_sections(primary: dict, fallback: dict) -> dict:
    merged = dict(primary)
    for field in ("components", "accessories", "sourceText"):
        if fallback.get(field) and (
            not merged.get(field)
            or (field != "sourceText" and weak_extracted_list(merged.get(field, "")))
        ):
            merged[field] = fallback[field]

    for prefix in ("component", "accessory"):
        status_field = f"{prefix}Status"
        keyword_field = f"{prefix}Keyword"
        if merged.get(status_field) == "not_found" and fallback.get(status_field) != "not_found":
            merged[status_field] = fallback.get(status_field, merged.get(status_field))
            merged[keyword_field] = fallback.get(keyword_field, merged.get(keyword_field))

    if merged.get("status") == "not_found" and fallback.get("status") != "not_found":
        merged["status"] = fallback["status"]
    if not merged.get("keyword") and fallback.get("keyword"):
        merged["keyword"] = fallback["keyword"]
    if primary.get("sourceText") and fallback.get("sourceText") and fallback["sourceText"] not in primary["sourceText"]:
        merged["sourceText"] = f"{primary['sourceText']}\n\nOCR来源：{fallback['sourceText']}"
    return merged


def choose_sections(page_sections: dict, manual_sections_result: dict) -> dict:
    if sections_useful(page_sections):
        return page_sections
    if sections_useful(manual_sections_result):
        return merge_sections(manual_sections_result, page_sections)
    return page_sections if page_sections.get("status") == "found" else empty_sections()


def build_payload() -> dict:
    pdftoppm = bundled_tool("pdftoppm")
    ocr_tool = compile_ocr_tool()
    manuals = load_manuals()
    model_types = load_model_types()
    skipped = {str(manual.get("id")): skip_reason(manual, model_types) for manual in manuals}
    skipped = {manual_id: reason for manual_id, reason in skipped.items() if reason}
    processed_manuals = [manual for manual in manuals if str(manual.get("id")) not in skipped]
    packing: dict[str, dict] = {}

    for index, manual in enumerate(processed_manuals, 1):
        typed_models = manual_model_types(manual, model_types)
        pdf_path = pdf_path_for(manual)
        manual_sections_result = manual_sections(manual)
        pdf_pages = extract_page_texts(pdf_path, PAGE_SCAN_LIMIT)
        page_candidate = best_page(pdf_pages)
        page_sections = (
            sections_from_text(page_candidate.get("text") or "")
            if page_candidate
            else empty_sections()
        )
        if candidate_needs_ocr(page_candidate, page_sections):
            ocr_pages = ocr_front_pages(pdf_path, pdftoppm, ocr_tool, OCR_SCAN_LIMIT)
            page_candidate = choose_best_with_ocr(pdf_pages, ocr_pages)
            page_sections = (
                sections_from_text(page_candidate.get("text") or "")
                if page_candidate
                else page_sections
            )
        sections = choose_sections(page_sections, manual_sections_result)
        if page_candidate:
            sections = choose_sections(
                sections_from_text(page_candidate.get("text") or ""),
                manual_sections_result,
            )

        image_url = ""
        page_number = None
        page_source = ""
        page_text = ""
        page_keyword = ""
        if page_candidate:
            page_number = page_candidate["page"]
            page_source = page_candidate["source"]
            page_keyword = page_candidate.get("keyword") or ""
            page_text = compact_source(page_candidate.get("text") or "", 1400)
            image_url = render_screenshot(
                pdf_path,
                str(manual.get("id")),
                page_number,
                pdftoppm,
                page_candidate.get("image_path") or "",
            )

        status = sections["status"]
        if status == "found" and not image_url:
            status = "text_only"

        packing[str(manual.get("id"))] = {
            "id": manual.get("id"),
            "title": manual.get("title", ""),
            "filename": manual.get("filename", ""),
            "models": list(typed_models.keys()),
            "machineTypes": sorted({value for value in typed_models.values() if value}),
            "status": status,
            "keyword": page_keyword or sections["keyword"],
            "page": page_number,
            "pageSource": page_source,
            "imageUrls": [image_url] if image_url else [],
            "components": sections["components"],
            "accessories": sections["accessories"],
            "sourceText": sections["sourceText"] or page_text,
            "pageText": page_text,
            "componentStatus": sections["componentStatus"],
            "accessoryStatus": sections["accessoryStatus"],
            "componentKeyword": sections["componentKeyword"],
            "accessoryKeyword": sections["accessoryKeyword"],
        }

        print(
            f"[{index:03d}/{len(processed_manuals)}] {manual.get('id')} "
            f"{status} page={page_number or '-'} types={'/'.join(packing[str(manual.get('id'))]['machineTypes']) or '-'} "
            f"keyword={packing[str(manual.get('id'))]['keyword'] or '-'}",
            flush=True,
        )

    build = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "manualCount": len(manuals),
        "processedCount": len(processed_manuals),
        "skippedCount": len(manuals) - len(processed_manuals),
        "processedMachineTypes": sorted(PROCESS_MACHINE_TYPES),
        "skippedReasons": {
            reason: sum(1 for item in skipped.values() if item == reason)
            for reason in sorted(set(skipped.values()))
        },
        "foundCount": sum(1 for item in packing.values() if item["status"] == "found"),
        "textOnlyCount": sum(1 for item in packing.values() if item["status"] == "text_only"),
        "notFoundCount": sum(1 for item in packing.values() if item["status"] == "not_found"),
        "screenshotCount": sum(1 for item in packing.values() if item["imageUrls"]),
        "ocrAvailable": bool(ocr_tool),
    }
    return {"build": build, "packing": packing}


def main() -> None:
    PACKING_DIR.mkdir(parents=True, exist_ok=True)
    TMP_DIR.mkdir(parents=True, exist_ok=True)

    payload = build_payload()
    PACKING_DATA_PATH.write_text(
        "window.PACKING_BUILD = "
        + json.dumps(payload["build"], ensure_ascii=False, separators=(",", ":"))
        + ";\nwindow.PACKING_LISTS = "
        + json.dumps(payload["packing"], ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    print(json.dumps(payload["build"], ensure_ascii=False, indent=2))
    print(PACKING_DATA_PATH)


if __name__ == "__main__":
    main()
